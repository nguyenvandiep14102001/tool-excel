import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError
from typing import Tuple, List, Dict, Any
import os
import re
import tempfile
import math

class ExcelUtils:
    """Utility class for Excel operations"""
    
    @staticmethod
    def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
        """Sanitize sheet name for Excel compatibility"""
        if not name or not isinstance(name, str):
            return "Sheet1"
        
        # Remove invalid characters: \ / * ? : [ ]
        sanitized = re.sub(r'[\\/*?:\[\]]', '', name)
        
        # Trim to max length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length-3] + "..."
        
        # Ensure not empty and doesn't start/end with apostrophe
        sanitized = sanitized.strip().strip("'")
        if not sanitized:
            sanitized = "Sheet1"
        
        return sanitized
    
    @staticmethod
    def clean_value(value: Any) -> Any:
        """Clean a single value to be JSON serializable"""
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return ""  # Replace NaN and None with empty string
        elif isinstance(value, (int, float)):
            # Handle infinity and very large numbers
            if math.isinf(value) or math.isnan(value):
                return ""
            return value
        elif isinstance(value, str):
            # Remove control characters and normalize string
            cleaned = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
            return cleaned.strip()
        else:
            # For other types, convert to string and clean
            try:
                str_value = str(value)
                cleaned = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', str_value)
                return cleaned.strip()
            except:
                return ""
    
    @staticmethod
    def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Sanitize DataFrame to remove problematic characters and NaN values"""
        df_clean = df.copy()
        
        for col in df_clean.columns:
            # Clean each value in the column
            df_clean[col] = df_clean[col].apply(ExcelUtils.clean_value)
        
        return df_clean
    
    @staticmethod
    def dataframe_to_dict_safe(df: pd.DataFrame) -> List[Dict]:
        """Convert DataFrame to dictionary safely for JSON serialization"""
        try:
            # First sanitize the DataFrame
            df_clean = ExcelUtils.sanitize_dataframe(df)
            
            # Convert to dictionary records
            records = df_clean.to_dict('records')
            
            # Additional cleaning for each record
            safe_records = []
            for record in records:
                safe_record = {}
                for key, value in record.items():
                    safe_record[str(key)] = ExcelUtils.clean_value(value)
                safe_records.append(safe_record)
            
            return safe_records
        except Exception as e:
            print(f"Error converting DataFrame to dict: {e}")
            return []
    
    @staticmethod
    def read_excel(file_path: str) -> pd.DataFrame:
        """Read Excel file into DataFrame, support both .xls and .xlsx"""
        try:
            # Check file extension to use appropriate engine
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.xls':
                # For .xls files, try xlrd first, then fallback to openpyxl
                try:
                    df = pd.read_excel(file_path, engine='xlrd')
                except ImportError:
                    # If xlrd not available, try openpyxl
                    df = pd.read_excel(file_path, engine='openpyxl')
            elif file_ext in ['.xlsx', '.xlsm']:
                # Use openpyxl for .xlsx files
                df = pd.read_excel(file_path, engine='openpyxl')
            else:
                # Try with default engine
                df = pd.read_excel(file_path)
            
            # Clean the DataFrame after reading
            return ExcelUtils.sanitize_dataframe(df)
                
        except Exception as e:
            raise Exception(f"Không thể đọc file {file_path}: {str(e)}")
    
    @staticmethod
    def save_excel_safe(df: pd.DataFrame, output_path: str, sheet_name: str = 'Result') -> str:
        """Safe method to save DataFrame to Excel - ALWAYS use .xlsx format"""
        try:
            # Clean the DataFrame first
            df_clean = ExcelUtils.sanitize_dataframe(df)
            safe_sheet_name = ExcelUtils.sanitize_sheet_name(sheet_name)
            
            # ALWAYS save as .xlsx to avoid xlwt dependency
            if not output_path.lower().endswith('.xlsx'):
                output_path = os.path.splitext(output_path)[0] + '.xlsx'
            
            # Use openpyxl engine
            df_clean.to_excel(output_path, sheet_name=safe_sheet_name, index=False, engine='openpyxl')
            
            return output_path
        except Exception as e:
            raise Exception(f"Lỗi khi lưu file: {str(e)}")
    
    @staticmethod
    def save_styled_excel(df: pd.DataFrame, output_path: str, styles: Dict[int, str] = None, 
                         sheet_name: str = 'Result') -> str:
        """Save DataFrame to Excel with optional styling"""
        try:
            # ALWAYS save as .xlsx for styling
            if not output_path.lower().endswith('.xlsx'):
                output_path = os.path.splitext(output_path)[0] + '.xlsx'
            
            safe_sheet_name = ExcelUtils.sanitize_sheet_name(sheet_name)
            df_clean = ExcelUtils.sanitize_dataframe(df)
            
            # First, create the Excel file without styling
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_clean.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            
            # Then apply styling if needed
            if styles:
                try:
                    wb = openpyxl.load_workbook(output_path)
                    ws = wb[safe_sheet_name]
                    
                    fills = {
                        'green': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
                        'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                        'blue': PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
                        'red': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                        'orange': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    }
                    
                    for row_idx, color in styles.items():
                        if color in fills and row_idx <= len(df_clean) + 1:  # Check row bounds
                            for col in range(1, len(df_clean.columns) + 1):
                                try:
                                    cell = ws.cell(row=row_idx, column=col)
                                    cell.fill = fills[color]
                                except (IllegalCharacterError, Exception):
                                    # Skip styling for problematic cells
                                    continue
                    
                    wb.save(output_path)
                except Exception as style_error:
                    # If styling fails, keep the unstyled version
                    print(f"Styling failed but file saved: {style_error}")
            
            return output_path
        except Exception as e:
            # Final fallback: save without any styling
            try:
                print(f"Advanced save failed, using basic save: {e}")
                return ExcelUtils.save_excel_safe(df, output_path, sheet_name)
            except Exception as final_error:
                raise Exception(f"Lỗi nghiêm trọng khi lưu file: {str(final_error)}")
    
    @staticmethod
    def apply_colors_to_excel(input_path: str, output_path: str, styles: Dict[int, str]) -> str:
        """Apply colors to an existing Excel file without modifying data"""
        try:
            # Only works with .xlsx files
            if not input_path.lower().endswith('.xlsx'):
                # Convert .xls to .xlsx first
                df = ExcelUtils.read_excel(input_path)
                return ExcelUtils.save_styled_excel(df, output_path, styles)
            
            # For .xlsx files, apply styling
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active
            
            fills = {
                'green': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
                'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                'blue': PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
            }
            
            for row_idx, color in styles.items():
                if color in fills:
                    for col in range(1, ws.max_column + 1):
                        try:
                            cell = ws.cell(row=row_idx, column=col)
                            cell.fill = fills[color]
                        except Exception:
                            continue
            
            wb.save(output_path)
            return output_path
        except Exception as e:
            # If styling fails, return original file
            import shutil
            shutil.copy2(input_path, output_path)
            return output_path

    @staticmethod
    def get_file_info(file_path: str) -> Dict[str, Any]:
        """Get basic file information with safe data handling"""
        if not os.path.exists(file_path):
            return {}
        
        try:
            df = ExcelUtils.read_excel(file_path)
            
            # Get safe sample data
            safe_sample_data = ExcelUtils.dataframe_to_dict_safe(df.head(3))
            
            return {
                'filename': os.path.basename(file_path),
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'file_extension': os.path.splitext(file_path)[1].lower(),
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': [str(col) for col in df.columns],
                'sample_data': safe_sample_data
            }
        except Exception as e:
            return {
                'filename': os.path.basename(file_path),
                'error': str(e)
            }
    
    @staticmethod
    def validate_excel_file(file_path: str) -> Tuple[bool, str]:
        """Validate if file is a readable Excel file"""
        if not os.path.exists(file_path):
            return False, "File không tồn tại"
        
        allowed_extensions = ['.xls', '.xlsx', '.xlsm']
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext not in allowed_extensions:
            return False, f"Định dạng file không được hỗ trợ: {file_ext}. Chỉ hỗ trợ: {', '.join(allowed_extensions)}"
        
        try:
            # Try to read the file to validate
            df = ExcelUtils.read_excel(file_path)
            if df.empty:
                return True, "File Excel hợp lệ nhưng không có dữ liệu"
            return True, "File Excel hợp lệ"
        except Exception as e:
            return False, f"Không thể đọc file Excel: {str(e)}"